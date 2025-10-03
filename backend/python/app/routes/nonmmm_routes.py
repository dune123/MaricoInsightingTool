"""
Non-MMM Analysis Routes
======================

Purpose: API endpoints specifically for Non-MMM analysis workflow

Description:
This module provides all the API endpoints needed for Non-MMM analysis:
1. Data summary statistics (mean, median, mode, std dev, etc.)
2. Histogram generation for data distribution visualization
3. Correlation matrix calculation
4. Variable type modification
5. Statistical analysis and insights
6. Data summary storage and retrieval for workflow persistence

Key Features:
- Statistical summary generation for all variables
- Histogram data generation for frontend charts
- Upper triangular correlation matrix
- Variable type modification with backend processing
- Data quality metrics and validation
- Data summary persistence for workflow continuity

Dependencies:
- FastAPI for API framework
- Pandas for data processing
- NumPy for statistical calculations
- Matplotlib/Seaborn for histogram generation
- Brand-specific file management
- JSON-based state persistence

Last Updated: 2025-01-31
Author: BrandBloom Backend Team
"""

from fastapi import APIRouter, HTTPException, Query, Body, File, UploadFile, Form
from typing import Dict, List, Any, Optional
import pandas as pd
import numpy as np
from datetime import datetime
import json
import logging
import os
import time
import io
from pathlib import Path
from openpyxl import load_workbook

from app.core.config import settings
from app.models.data_models import BaseResponse, ErrorResponse
from app.utils.file_utils import find_file_with_fallback

# Configure logging
logger = logging.getLogger(__name__)

def detect_excel_data_type(file_path: Path, column_name: str, col_data: pd.Series) -> str:
    """
    Detect data type from Excel file by reading actual cell formats
    
    Args:
        file_path: Path to Excel file
        column_name: Name of the column
        col_data: Pandas Series with the column data
        
    Returns:
        str: Detected data type ('numeric', 'datetime', 'percentage', 'character')
    """
    try:
        # First check basic pandas data types
        if pd.api.types.is_datetime64_any_dtype(col_data):
            return "datetime"
        elif not pd.api.types.is_numeric_dtype(col_data):
            return "character"
        
        # For numeric data, check Excel cell formats
        try:
            # Load the Excel file with openpyxl to read cell formats
            workbook = load_workbook(file_path, data_only=False)
            
            # Try to find the sheet and column
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Find the column by looking at the first row (headers)
                column_index = None
                for col_idx, cell in enumerate(sheet[1], 1):  # First row
                    if cell.value == column_name:
                        column_index = col_idx
                        break
                
                if column_index is not None:
                    # Check a few cells in this column to determine format
                    percentage_count = 0
                    total_checked = 0
                    
                    # Check first 10 non-empty cells in the column
                    for row_idx in range(2, min(12, sheet.max_row + 1)):  # Skip header row
                        cell = sheet.cell(row=row_idx, column=column_index)
                        if cell.value is not None and str(cell.value).strip():
                            total_checked += 1
                            
                            # Check if the cell format is percentage
                            if cell.number_format and '%' in str(cell.number_format):
                                percentage_count += 1
                    
                    # If more than 50% of checked cells have percentage format, mark as percentage
                    if total_checked > 0 and (percentage_count / total_checked) > 0.5:
                        return "percentage"
                    
                    break  # Found the column, no need to check other sheets
            
            workbook.close()
            
        except Exception as e:
            logger.warning(f"Could not read Excel formats for {column_name}: {e}")
        
        # Fallback to numeric if no percentage format detected
        return "numeric"
        
    except Exception as e:
        logger.error(f"Error detecting data type for {column_name}: {e}")
        # Fallback to basic pandas detection
        if pd.api.types.is_numeric_dtype(col_data):
            return "numeric"
        elif pd.api.types.is_datetime64_any_dtype(col_data):
            return "datetime"
        else:
            return "character"

router = APIRouter(prefix="/api/nonmmm", tags=["Non-MMM Analysis"])

def convert_numpy_types(obj):
    """
    Recursively convert numpy types to Python native types to prevent serialization errors
    """
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {key: convert_numpy_types(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_numpy_types(item) for item in obj]
    else:
        return obj

# ========================================
# DATA SUMMARY STORAGE AND RETRIEVAL
# ========================================

@router.post("/store-summary")
async def store_data_summary(
    request_data: Dict[str, Any] = Body(...)
) -> Dict[str, Any]:
    """
    Store data summary for Non-MMM analysis workflow persistence
    
    Request body should contain:
    - analysisId: Unique analysis identifier
    - brand: Brand name for data organization
    - filename: Name of the analyzed file
    - dataSummary: Complete data summary object
    """
    try:
        analysis_id = request_data.get("analysisId")
        brand = request_data.get("brand")
        filename = request_data.get("filename")
        data_summary = request_data.get("dataSummary")
        
        if not all([analysis_id, brand, filename, data_summary]):
            raise HTTPException(
                status_code=400, 
                detail="analysisId, brand, filename, and dataSummary are required"
            )
        
        # Get brand directories and create them if they don't exist
        brand_dirs = settings.get_brand_directories(brand)
        settings.create_brand_directories(brand)
        
        # Create metadata directory for non-MMM summaries
        summary_dir = brand_dirs["metadata_dir"] / "nonmmm_summaries"
        summary_dir.mkdir(parents=True, exist_ok=True)
        
        # Create summary file path
        summary_file = summary_dir / f"{analysis_id}.json"
        
        # Prepare summary data with metadata
        summary_data = {
            "analysisId": analysis_id,
            "brand": brand,
            "filename": filename,
            "dataSummary": data_summary,
            "storedAt": datetime.now().isoformat(),
            "version": "1.0"
        }
        
        # Save to JSON file
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(summary_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"✅ Data summary stored successfully for analysis: {analysis_id}")
        
        return {
            "success": True,
            "message": f"Data summary stored successfully for analysis {analysis_id}",
            "data": {
                "analysisId": analysis_id,
                "storedAt": summary_data["storedAt"],
                "filePath": str(summary_file)
            }
        }
        
    except Exception as e:
        logger.error(f"Error storing data summary: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to store data summary: {str(e)}"
        )

@router.get("/get-summary/{analysis_id}")
async def get_stored_data_summary(
    analysis_id: str,
    brand: str = Query(..., description="Brand name for data lookup")
) -> Dict[str, Any]:
    """
    Retrieve stored data summary for Non-MMM analysis workflow
    
    Returns the complete data summary that was previously stored
    """
    try:
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Check if brand directories exist
        if not brand_dirs["metadata_dir"].exists():
            raise HTTPException(
                status_code=404, 
                detail=f"Brand {brand} not found or no data uploaded"
            )
        
        # Look for summary file
        summary_dir = brand_dirs["metadata_dir"] / "nonmmm_summaries"
        summary_file = summary_dir / f"{analysis_id}.json"
        
        if not summary_file.exists():
            raise HTTPException(
                status_code=404, 
                detail=f"Data summary not found for analysis {analysis_id}"
            )
        
        # Read and return summary data
        with open(summary_file, 'r', encoding='utf-8') as f:
            summary_data = json.load(f)
        
        logger.info(f"✅ Stored data summary retrieved successfully for analysis: {analysis_id}")
        
        return {
            "success": True,
            "message": f"Data summary retrieved successfully for analysis {analysis_id}",
            "data": summary_data
        }
        
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        logger.error(f"Error retrieving stored data summary: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to retrieve stored data summary: {str(e)}"
        )

# ========================================
# TEST ENDPOINTS (for debugging)
# ========================================

@router.get("/test-cors")
async def test_cors() -> Dict[str, Any]:
    """
    Simple test endpoint to verify CORS is working
    
    This helps isolate CORS issues from other endpoint problems
    """
    print("🔍 DEBUG: test-cors endpoint called successfully")
    return {
        "success": True,
        "message": "CORS test endpoint working",
        "timestamp": datetime.now().isoformat(),
        "cors_test": "If you see this, CORS is working"
    }

@router.post("/generate-powerpoint")
async def generate_powerpoint_presentation(
    analysis_id: str,
    brand: str,
    analysis_data: Dict[str, Any]
):
    """
    Generate PowerPoint presentation for Non-MMM analysis
    
    Args:
        analysis_id: Unique analysis identifier
        brand: Brand name for the analysis
        analysis_data: Complete analysis data including charts, models, etc.
        
    Returns:
        FileResponse with the generated PowerPoint file
    """
    try:
        from app.services.powerpoint_service import PowerPointService
        from fastapi.responses import FileResponse
        
        # Validate required parameters
        if not analysis_id or not brand:
            raise HTTPException(
                status_code=400,
                detail="analysis_id and brand are required parameters"
            )
        
        # Generate PowerPoint presentation
        output_path = PowerPointService.generate_analysis_presentation(
            analysis_id=analysis_id,
            brand_name=brand,
            analysis_data=analysis_data
        )
        
        # Return the file for download
        return FileResponse(
            path=str(output_path),
            filename=output_path.name,
            media_type='application/vnd.openxmlformats-officedocument.presentationml.presentation'
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate PowerPoint presentation: {str(e)}"
        )

# ========================================
# DATA SUMMARY ENDPOINTS
# ========================================

@router.get("/data-summary/{filename}")
async def get_data_summary(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup")
) -> Dict[str, Any]:
    """
    Get comprehensive statistical summary for all variables in the dataset
    
    Returns:
    - Variable names and types
    - Statistical measures (mean, median, mode, std dev, min, max)
    - Data quality metrics (null count, unique count)
    - Descriptive statistics for each variable
    """
    try:
        # Get brand directories and create them if they don't exist
        brand_dirs = settings.get_brand_directories(brand)
        
        # Create brand directories if they don't exist
        settings.create_brand_directories(brand)
        
        # CRITICAL: Always read from RAW file to ensure consistency with delete operations
        # Data summary should show the same data that gets modified by delete operations
        search_directories = [
            brand_dirs["raw_dir"],  # Read from raw file first
            settings.UPLOAD_DIR / "raw",  # Global upload directory as fallback
            settings.RAW_DIR  # Legacy raw directory as fallback
        ]
        file_path, source_dir = find_file_with_fallback(filename, search_directories)
        
        # Log which file we're reading for debugging
        logger.info(f"Reading data summary from file: {file_path} (found in: {source_dir})")
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Read the data
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        # AUTOMATIC NULL COLUMN DELETION: Remove columns with >50% null values
        # This ensures data quality for analysis
        total_rows = len(df)
        null_threshold = 0.5  # 50% threshold
        columns_to_delete = []
        
        for column in df.columns:
            null_count = df[column].isnull().sum()
            null_percentage = null_count / total_rows if total_rows > 0 else 0
            
            if null_percentage > null_threshold:
                columns_to_delete.append(column)
                logger.info(f"Auto-deleting column '{column}' with {null_percentage:.1%} null values")
        
        # Delete columns with high null values
        if columns_to_delete:
            df = df.drop(columns=columns_to_delete)
            # Save the cleaned data back to the file
            if filename.endswith('.csv'):
                df.to_csv(file_path, index=False)
            else:
                df.to_excel(file_path, index=False)
            
            logger.info(f"Auto-deleted {len(columns_to_delete)} columns with >50% null values: {columns_to_delete}")
        
        # Generate summary for each variable
        summary_data = []
        
        # Check if we have stored column type preferences for this file
        stored_types_file = brand_dirs["metadata_dir"] / f"{filename}_column_types.json"
        stored_column_types = {}
        
        if stored_types_file.exists():
            try:
                with open(stored_types_file, 'r', encoding='utf-8') as f:
                    stored_column_types = json.load(f)
                logger.info(f"Loaded stored column types for {filename}: {stored_column_types}")
            except Exception as e:
                logger.warning(f"Failed to load stored column types: {e}")
        
        for column in df.columns:
            col_data = df[column]
            
            # Use stored type if available, otherwise detect from data
            if column in stored_column_types:
                var_type = stored_column_types[column]
                logger.info(f"Using stored type for {column}: {var_type}")
            else:
                # Determine data type from actual data and Excel cell formats
                var_type = detect_excel_data_type(file_path, column, col_data)
                logger.info(f"Detected type for {column}: {var_type}")
            
            # Basic statistics
            variable_summary = {
                "name": column,
                "type": var_type,
                "count": len(col_data.dropna()),
                "nullCount": col_data.isnull().sum(),
                "uniqueCount": col_data.nunique()
            }
            
            # Type-specific statistics
            # Note: percentage type is treated the same as numeric since percentage values are numeric
            if var_type == "numeric" or var_type == "percentage":
                try:
                    # Basic statistical measures
                    mean_val = col_data.mean()
                    median_val = col_data.median()
                    mode_val = col_data.mode()
                    min_val = col_data.min()
                    max_val = col_data.max()
                    std_val = col_data.std()
                    var_val = col_data.var()
                    skew_val = col_data.skew()
                    kurt_val = col_data.kurtosis()
                    
                    # Calculate P6M (Past 6 months average) and MAT (Moving Annual Total)
                    p6m_val = None
                    mat_val = None
                    
                    # If we have a datetime column (Month), calculate time-based metrics
                    if 'Month' in df.columns and pd.api.types.is_datetime64_any_dtype(df['Month']):
                        try:
                            # Sort by date to ensure chronological order
                            sorted_df = df.sort_values('Month').reset_index(drop=True)
                            sorted_col = sorted_df[column]
                            
                            # Calculate P6M (last 6 months if available)
                            if len(sorted_col) >= 6:
                                p6m_val = float(sorted_col.tail(6).mean())
                            
                            # Calculate MAT (last 12 months if available)
                            if len(sorted_col) >= 12:
                                mat_val = float(sorted_col.tail(12).mean())
                            elif len(sorted_col) >= 6:
                                # If less than 12 months, use available months
                                mat_val = float(sorted_col.mean())
                        except Exception as time_error:
                            print(f"⚠️ WARNING: Failed to calculate time-based metrics for {column}: {time_error}")
                    
                    variable_summary.update({
                        "mean": float(mean_val) if pd.notna(mean_val) else None,
                        "median": float(median_val) if pd.notna(median_val) else None,
                        "mode": float(mode_val.iloc[0]) if len(mode_val) > 0 else None,
                        "min": float(min_val) if pd.notna(min_val) else None,
                        "max": float(max_val) if pd.notna(max_val) else None,
                        "stdDev": float(std_val) if pd.notna(std_val) else None,
                        "variance": float(var_val) if pd.notna(var_val) else None,
                        "skewness": float(skew_val) if pd.notna(skew_val) else None,
                        "kurtosis": float(kurt_val) if pd.notna(kurt_val) else None,
                        "p6m": p6m_val,  # Past 6 months average
                        "mat": mat_val    # Moving Annual Total
                    })
                except Exception as e:
                    print(f"❌ ERROR: Failed to calculate numeric stats for '{column}': {e}")
                    # Set all numeric stats to None on error
                    variable_summary.update({
                        "mean": None, "median": None, "mode": None, "min": None, 
                        "max": None, "stdDev": None, "variance": None, 
                        "skewness": None, "kurtosis": None, "p6m": None, "mat": None
                    })
            elif var_type == "datetime":
                variable_summary.update({
                    "min": col_data.min().timestamp() * 1000 if pd.notna(col_data.min()) else None,  # Convert to JS timestamp
                    "max": col_data.max().timestamp() * 1000 if pd.notna(col_data.max()) else None,
                    "mode": col_data.mode().iloc[0].strftime('%Y-%m-%d') if len(col_data.mode()) > 0 else None
                })
            else:  # character
                mode_val = col_data.mode().iloc[0] if len(col_data.mode()) > 0 else None
                variable_summary.update({
                    "mode": str(mode_val) if mode_val is not None else None
                })
            
            summary_data.append(variable_summary)
        
        # Convert any remaining numpy types to Python native types
        response_data = {
            "success": True,
            "message": f"Data summary generated for {len(summary_data)} variables",
            "data": {
                "variables": summary_data,
                "totalRows": int(len(df)),
                "totalColumns": int(len(df.columns)),
                "filename": filename,
                "generatedAt": datetime.now().isoformat()
            }
        }
        
        # Apply numpy type conversion to the entire response
        return convert_numpy_types(response_data)
        
    except Exception as e:
        logger.error(f"Error generating data summary: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate data summary: {str(e)}")

@router.post("/modify-column-type/{filename}")
async def modify_column_type(
    filename: str,
    request_data: Dict[str, Any] = Body(...),
    brand: str = Query(..., description="Brand name for data lookup")
) -> Dict[str, Any]:
    """
    Modify the data type of a specific column
    
    Request body should contain:
    - columnName: Name of the column to modify
    - newType: New data type (numeric, datetime, percentage, character)
    """
    try:
        column_name = request_data.get("columnName")
        new_type = request_data.get("newType")
        
        if not column_name or not new_type:
            raise HTTPException(status_code=400, detail="columnName and newType are required")
        
        # Get brand directories and create them if they don't exist
        brand_dirs = settings.get_brand_directories(brand)
        
        # Create brand directories if they don't exist
        settings.create_brand_directories(brand)
        
        # Find the file
        search_directories = [brand_dirs["concat_dir"], brand_dirs["raw_dir"]]
        file_path, _ = find_file_with_fallback(filename, search_directories)
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Read the data
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        if column_name not in df.columns:
            raise HTTPException(status_code=404, detail=f"Column {column_name} not found in dataset")
        
        # Convert column type
        original_type = str(df[column_name].dtype)
        
        try:
            if new_type == "numeric":
                df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
            elif new_type == "datetime":
                df[column_name] = pd.to_datetime(df[column_name], errors='coerce')
            elif new_type == "percentage":
                # For percentage, we keep the original values but mark the type
                # The division by 100 was causing data loss - we just want to mark it as percentage
                df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
            elif new_type == "character":
                df[column_name] = df[column_name].astype(str)
            else:
                raise HTTPException(status_code=400, detail=f"Unsupported type: {new_type}")
            
            # Save the modified file back
            if filename.endswith('.csv'):
                df.to_csv(file_path, index=False)
            else:
                df.to_excel(file_path, index=False)
            
            # Save the column type preference to metadata
            try:
                stored_types_file = brand_dirs["metadata_dir"] / f"{filename}_column_types.json"
                stored_column_types = {}
                
                if stored_types_file.exists():
                    with open(stored_types_file, 'r', encoding='utf-8') as f:
                        stored_column_types = json.load(f)
                
                # Update the stored type for this column
                stored_column_types[column_name] = new_type
                
                # Save the updated preferences
                with open(stored_types_file, 'w', encoding='utf-8') as f:
                    json.dump(stored_column_types, f, indent=2, ensure_ascii=False)
                
                logger.info(f"Saved column type preference for {column_name}: {new_type}")
                
            except Exception as save_error:
                logger.warning(f"Failed to save column type preference: {save_error}")
                # Don't fail the operation if metadata saving fails
            
            return {
                "success": True,
                "message": f"Column {column_name} type changed from {original_type} to {new_type}",
                "data": {
                    "columnName": column_name,
                    "originalType": original_type,
                    "newType": new_type,
                    "modifiedAt": datetime.now().isoformat()
                }
            }
            
        except Exception as conv_error:
            raise HTTPException(status_code=400, detail=f"Failed to convert column to {new_type}: {str(conv_error)}")
        
    except Exception as e:
        logger.error(f"Error modifying column type: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to modify column type: {str(e)}")

# ========================================
# HISTOGRAM GENERATION ENDPOINTS
# ========================================

@router.get("/histograms/{filename}")
async def get_histograms(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup"),
    bins: str = Query("10", description="Number of bins for histograms or 'auto' for intelligent binning")
) -> Dict[str, Any]:
    """
    Generate histogram data for all numeric variables in the dataset
    
    Returns histogram data that can be used by frontend charting libraries
    """
    try:
        # Get brand directories and create them if they don't exist
        brand_dirs = settings.get_brand_directories(brand)
        
        # Create brand directories if they don't exist
        settings.create_brand_directories(brand)
        
        # CRITICAL: Always read from RAW file to ensure consistency with delete operations
        # Histograms should show the same data that gets modified by delete operations
        search_directories = [brand_dirs["raw_dir"]]
        file_path, source_dir = find_file_with_fallback(filename, search_directories)
        
        # Log which file we're reading for debugging
        logger.info(f"Reading histogram data from file: {file_path} (found in: {source_dir})")
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Read the data
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        # AUTOMATIC NULL COLUMN DELETION: Remove columns with >50% null values
        # This ensures data quality for analysis
        total_rows = len(df)
        null_threshold = 0.5  # 50% threshold
        columns_to_delete = []
        
        for column in df.columns:
            null_count = df[column].isnull().sum()
            null_percentage = null_count / total_rows if total_rows > 0 else 0
            
            if null_percentage > null_threshold:
                columns_to_delete.append(column)
                logger.info(f"Auto-deleting column '{column}' with {null_percentage:.1%} null values")
        
        # Delete columns with high null values
        if columns_to_delete:
            df = df.drop(columns=columns_to_delete)
            # Save the cleaned data back to the file
            if filename.endswith('.csv'):
                df.to_csv(file_path, index=False)
            else:
                df.to_excel(file_path, index=False)
            
            logger.info(f"Auto-deleted {len(columns_to_delete)} columns with >50% null values: {columns_to_delete}")
        
        histograms = []
        
        # Generate histograms for numeric columns
        for column in df.columns:
            if pd.api.types.is_numeric_dtype(df[column]):
                col_data = df[column].dropna()
                
                if len(col_data) > 0:
                    # Intelligent binning
                    if bins == "auto":
                        # Use Sturges' formula for optimal bin count
                        n_bins = int(np.ceil(1 + 3.322 * np.log10(len(col_data))))
                        # Ensure reasonable range
                        n_bins = max(5, min(n_bins, 20))
                    else:
                        try:
                            n_bins = int(bins)
                            n_bins = max(3, min(n_bins, 30))  # Reasonable bounds
                        except ValueError:
                            n_bins = 10
                    
                    # Generate histogram
                    counts, bin_edges = np.histogram(col_data, bins=n_bins)
                    
                    # Prepare data for frontend
                    from app.utils.number_formatter import format_histogram_range_label
                    
                    histogram_data = {
                        "variableName": column,
                        "bins": bin_edges.tolist(),
                        "counts": counts.tolist(),
                        "binWidth": float(bin_edges[1] - bin_edges[0]) if len(bin_edges) > 1 else 0,
                        "binCount": n_bins,
                        "chartData": {
                            "labels": [format_histogram_range_label(bin_edges[i], bin_edges[i+1]) for i in range(len(counts))],
                            "datasets": [{
                                "label": column,
                                "data": counts.tolist(),
                                "backgroundColor": "rgba(59, 130, 246, 0.6)",
                                "borderColor": "rgba(59, 130, 246, 1)",
                                "borderWidth": 1
                            }]
                        }
                    }
                    
                    histograms.append(histogram_data)
        
        return {
            "success": True,
            "message": f"Generated histograms for {len(histograms)} numeric variables",
            "data": {
                "histograms": histograms,
                "totalVariables": len(histograms),
                "filename": filename,
                "generatedAt": datetime.now().isoformat()
            }
        }
        
    except Exception as e:
        logger.error(f"Error generating histograms: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate histograms: {str(e)}")

# ========================================
# CORRELATION MATRIX ENDPOINTS
# ========================================

@router.get("/correlation-matrix/{filename}")
async def get_correlation_matrix(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup"),
    method: str = Query("pearson", description="Correlation method: pearson, spearman, or kendall"),
    variables: str = Query(None, description="Comma-separated list of variables to include in correlation matrix")
) -> Dict[str, Any]:
    """
    Generate correlation matrix for all numeric variables
    
    Returns upper triangular correlation matrix for visualization
    """
    try:
        # Get brand directories and create them if they don't exist
        brand_dirs = settings.get_brand_directories(brand)
        
        # Create brand directories if they don't exist
        settings.create_brand_directories(brand)
        
        # Find the file
        search_directories = [brand_dirs["concat_dir"], brand_dirs["raw_dir"]]
        file_path, _ = find_file_with_fallback(filename, search_directories)
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Read the data
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        # Select only numeric columns
        numeric_df = df.select_dtypes(include=[np.number])
        
        if numeric_df.empty:
            return {
                "success": False,
                "message": "No numeric variables found for correlation analysis",
                "data": None
            }
        
        # Filter to selected variables if provided
        if variables:
            try:
                selected_vars = [var.strip() for var in variables.split(',') if var.strip()]
                # Only include variables that exist in the dataframe and are numeric
                available_vars = [var for var in selected_vars if var in numeric_df.columns]
                
                if not available_vars:
                    return {
                        "success": False,
                        "message": f"None of the selected variables ({', '.join(selected_vars)}) are numeric or found in the dataset",
                        "data": None
                    }
                
                numeric_df = numeric_df[available_vars]
                logger.info(f"Filtered correlation matrix to selected variables: {available_vars}")
                
            except Exception as e:
                logger.warning(f"Error parsing variables parameter: {str(e)}, using all numeric variables")
        
        # Calculate correlation matrix
        corr_matrix = numeric_df.corr(method=method)
        
        # Convert to lists for JSON serialization
        variables = corr_matrix.columns.tolist()
        correlations = corr_matrix.values.tolist()
        
        # Replace NaN values with None for JSON serialization
        for i in range(len(correlations)):
            for j in range(len(correlations[i])):
                if pd.isna(correlations[i][j]):
                    correlations[i][j] = None
        
        return {
            "success": True,
            "message": f"Generated correlation matrix for {len(variables)} variables using {method} method",
            "data": {
                "variables": variables,
                "correlations": correlations,
                "method": method,
                "filename": filename,
                "generatedAt": datetime.now().isoformat()
            }
        }
        
    except Exception as e:
        logger.error(f"Error generating correlation matrix: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate correlation matrix: {str(e)}")

# ========================================
# DATA VALIDATION ENDPOINTS
# ========================================

@router.get("/data-validation/{filename}")
async def validate_data_quality(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup")
) -> Dict[str, Any]:
    """
    Perform comprehensive data quality validation
    
    Returns data quality metrics and recommendations
    """
    try:
        # Get brand directories and create them if they don't exist
        brand_dirs = settings.get_brand_directories(brand)
        
        # Create brand directories if they don't exist
        settings.create_brand_directories(brand)
        
        # Find the file
        search_directories = [brand_dirs["concat_dir"], brand_dirs["raw_dir"]]
        file_path, _ = find_file_with_fallback(filename, search_directories)
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Read the data
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        validation_results = {
            "totalRows": len(df),
            "totalColumns": len(df.columns),
            "dataQuality": "good",  # Default
            "issues": [],
            "recommendations": [],
            "columnIssues": []
        }
        
        # Check for missing data
        total_cells = df.shape[0] * df.shape[1]
        missing_cells = df.isnull().sum().sum()
        missing_percentage = (missing_cells / total_cells) * 100
        
        if missing_percentage > 20:
            validation_results["issues"].append(f"High missing data: {missing_percentage:.1f}% of cells are empty")
            validation_results["dataQuality"] = "poor"
        elif missing_percentage > 10:
            validation_results["issues"].append(f"Moderate missing data: {missing_percentage:.1f}% of cells are empty")
            validation_results["dataQuality"] = "fair"
        
        # Check individual columns
        for column in df.columns:
            col_missing = (df[column].isnull().sum() / len(df)) * 100
            col_unique = df[column].nunique()
            
            column_issue = {
                "columnName": column,
                "missingPercentage": col_missing,
                "uniqueValues": col_unique,
                "issues": []
            }
            
            if col_missing > 50:
                column_issue["issues"].append("Very high missing data")
            elif col_missing > 25:
                column_issue["issues"].append("High missing data")
            
            if col_unique == 1:
                column_issue["issues"].append("All values are the same")
            elif col_unique == len(df):
                column_issue["issues"].append("All values are unique")
            
            if column_issue["issues"]:
                validation_results["columnIssues"].append(column_issue)
        
        # Generate recommendations
        if missing_percentage > 10:
            validation_results["recommendations"].append("Consider data imputation or removal of high-missing columns")
        
        if len(validation_results["columnIssues"]) > 0:
            validation_results["recommendations"].append("Review flagged columns for data quality issues")
        
        return {
            "success": True,
            "message": "Data validation completed",
            "data": validation_results
        }
        
    except Exception as e:
        logger.error(f"Error validating data quality: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to validate data quality: {str(e)}")

# ========================================
# CHART ANALYSIS ENDPOINTS
# ========================================

@router.get("/chart-data/{filename}")
async def get_chart_data(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup"),
    target_variable: str = Query(..., description="Target variable name"),
    trendline_type: str = Query("linear", description="Trendline type: linear, polynomial-2, polynomial-3")
) -> Dict[str, Any]:
    """
    Generate chart data for Non-MMM analysis including line charts and scatter plots with trendlines
    
    Returns chart data for all variables against the target variable:
    - Line charts: Target variable vs time (month) with trendline
    - Scatter plots: Target variable vs independent variable with trendline
    - Expected vs unexpected result classification based on trendline slope
    """
    try:
        # Get brand directories and create them if they don't exist
        brand_dirs = settings.get_brand_directories(brand)
        settings.create_brand_directories(brand)
        
        # CRITICAL: Always read from RAW file to ensure consistency with delete operations
        # Charts should show the same data that gets modified by delete operations
        search_directories = [brand_dirs["raw_dir"]]
        file_path, source_dir = find_file_with_fallback(filename, search_directories)
        
        # Log which file we're reading for debugging
        logger.info(f"Reading chart data from file: {file_path} (found in: {source_dir})")
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Read the data
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        if target_variable not in df.columns:
            raise HTTPException(status_code=404, detail=f"Target variable {target_variable} not found in dataset")
        
        # AUTOMATIC NULL COLUMN DELETION: Remove columns with >50% null values
        # This ensures data quality for charting analysis
        total_rows = len(df)
        null_threshold = 0.5  # 50% threshold
        columns_to_delete = []
        
        for column in df.columns:
            if column != target_variable:  # Don't delete target variable
                null_count = df[column].isnull().sum()
                null_percentage = null_count / total_rows if total_rows > 0 else 0
                
                if null_percentage > null_threshold:
                    columns_to_delete.append(column)
                    logger.info(f"Auto-deleting column '{column}' with {null_percentage:.1%} null values")
        
        # Delete columns with high null values
        if columns_to_delete:
            df = df.drop(columns=columns_to_delete)
            # Save the cleaned data back to the file
            if filename.endswith('.csv'):
                df.to_csv(file_path, index=False)
            else:
                df.to_excel(file_path, index=False)
            
            logger.info(f"Auto-deleted {len(columns_to_delete)} columns with >50% null values: {columns_to_delete}")
        
        # Update target variable check after potential deletions
        if target_variable not in df.columns:
            raise HTTPException(status_code=404, detail=f"Target variable {target_variable} not found in dataset after null column cleanup")
        
        # Get all numeric columns except target variable
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        independent_variables = [col for col in numeric_columns if col != target_variable]
        
        # Look for time/month column
        time_column = None
        for col in df.columns:
            if 'month' in col.lower() or 'date' in col.lower() or 'time' in col.lower():
                time_column = col
                break
        
        chart_data = []
        
        for variable in independent_variables:
            # Prepare data for this variable - don't drop NaN for line chart to preserve gaps
            scatter_df = df[[target_variable, variable]].dropna()
            
            if len(scatter_df) < 3:  # Need at least 3 points for trendline
                continue
            
            x_values = scatter_df[variable].values
            y_values = scatter_df[target_variable].values
            
            # Calculate trendline coefficients
            trendline_data = _calculate_trendline(x_values, y_values, trendline_type)
            
            # Generate scatter plot data (target variable vs independent variable)
            scatter_data = {
                "x": x_values.tolist(),
                "y": y_values.tolist(),
                "trendline": trendline_data["trendline_points"],
                "slope": trendline_data["slope"],
                "r_squared": trendline_data["r_squared"]
            }
            
            # Generate line chart data (if time column exists)
            line_data = None
            if time_column and time_column in df.columns:
                # Don't drop NaN for line chart to preserve gaps in the line
                time_df = df[[target_variable, variable, time_column]].copy()
                if len(time_df) > 0:
                    # Sort by time
                    time_df = time_df.sort_values(time_column)
                    
                    # Convert dates to MMM-YY format
                    try:
                        time_df[time_column] = pd.to_datetime(time_df[time_column])
                        formatted_dates = time_df[time_column].dt.strftime('%b-%y').tolist()
                    except:
                        # Fallback to string format if date conversion fails
                        formatted_dates = time_df[time_column].astype(str).tolist()
                    
                    line_data = {
                        "time": formatted_dates,
                        "target_values": time_df[target_variable].tolist(),
                        "variable_values": time_df[variable].tolist()
                    }
            
            chart_data.append({
                "variable": variable,
                "scatter_plot": scatter_data,
                "line_chart": line_data,
                "trendline_type": trendline_type
            })
        
        return {
            "success": True,
            "message": f"Generated chart data for {len(chart_data)} variables",
            "data": {
                "charts": chart_data,
                "target_variable": target_variable,
                "time_column": time_column,
                "total_variables": len(chart_data),
                "filename": filename,
                "generatedAt": datetime.now().isoformat()
            }
        }
        
    except Exception as e:
        logger.error(f"Error generating chart data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate chart data: {str(e)}")

def _calculate_trendline(x_values, y_values, trendline_type):
    """
    Calculate trendline coefficients and generate trendline points
    """
    try:
        from sklearn.preprocessing import PolynomialFeatures
        from sklearn.linear_model import LinearRegression
        from sklearn.pipeline import Pipeline
        from sklearn.metrics import r2_score
        
        # Reshape data for sklearn
        X = x_values.reshape(-1, 1)
        y = y_values
        
        # Create model based on trendline type
        if trendline_type == "polynomial-2":
            model = Pipeline([
                ('poly', PolynomialFeatures(degree=2)),
                ('linear', LinearRegression())
            ])
        elif trendline_type == "polynomial-3":
            model = Pipeline([
                ('poly', PolynomialFeatures(degree=3)),
                ('linear', LinearRegression())
            ])
        else:  # linear
            model = LinearRegression()
        
        # Fit the model
        model.fit(X, y)
        
        # Generate trendline points
        x_min, x_max = x_values.min(), x_values.max()
        x_range = np.linspace(x_min, x_max, 50).reshape(-1, 1)
        y_pred = model.predict(x_range)
        
        # Calculate R-squared
        y_pred_original = model.predict(X)
        r_squared = r2_score(y, y_pred_original)
        
        # Calculate slope (for linear or approximate slope for polynomial)
        if trendline_type == "linear":
            slope = model.coef_[0] if hasattr(model, 'coef_') else 0
        else:
            # For polynomial, calculate average slope across the range
            slope = (y_pred[-1] - y_pred[0]) / (x_range[-1][0] - x_range[0][0])
        
        return {
            "trendline_points": {
                "x": x_range.flatten().tolist(),
                "y": y_pred.tolist()
            },
            "slope": float(slope),
            "r_squared": float(r_squared)
        }
        
    except Exception as e:
        logger.warning(f"Failed to calculate trendline: {e}")
        # Fallback to simple linear regression
        try:
            slope, intercept = np.polyfit(x_values, y_values, 1)
            x_min, x_max = x_values.min(), x_values.max()
            x_range = np.linspace(x_min, x_max, 50)
            y_range = slope * x_range + intercept
            
            return {
                "trendline_points": {
                    "x": x_range.tolist(),
                    "y": y_range.tolist()
                },
                "slope": float(slope),
                "r_squared": 0.0  # Unable to calculate without sklearn
            }
        except:
            return {
                "trendline_points": {"x": [], "y": []},
                "slope": 0.0,
                "r_squared": 0.0
            }

# ========================================
# STATISTICAL MODELING ENDPOINTS
# ========================================

@router.post("/train-model")
async def train_statistical_model(
    request_data: Dict[str, Any] = Body(...)
) -> Dict[str, Any]:
    """
    Train a statistical model for Non-MMM analysis
    
    Request body should contain:
    - filename: Name of the analyzed file
    - targetVariable: Target variable for prediction
    - independentVariables: List of independent variables
    - modelType: Type of model to train (linear, log-linear, log-log, ridge, bayesian)
    - modelParameters: Optional model-specific parameters
    - validationSplit: Optional validation split ratio (default: 0.2)
    - dataType: Data type to use ('original' or 'standardized', default: 'original')
    """
    try:
        filename = request_data.get("filename")
        target_variable = request_data.get("targetVariable")
        independent_variables = request_data.get("independentVariables", [])
        model_type = request_data.get("modelType", "linear")
        model_parameters = request_data.get("modelParameters", {})
        validation_split = request_data.get("validationSplit", 0.2)
        data_type = request_data.get("dataType", "original")  # 'original' or 'standardized'
        
        if not all([filename, target_variable, independent_variables]):
            raise HTTPException(
                status_code=400,
                detail="filename, targetVariable, and independentVariables are required"
            )
        
        # Get brand from request data
        brand = request_data.get("brand")
        if not brand:
            raise HTTPException(status_code=400, detail="Brand parameter is required")
        
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Find the file based on data type
        if data_type == "standardized":
            # STRICT: Only use standardized data, NO FALLBACKS
            from app.services.data_standardization_service import DataStandardizationService
            
            # First try to find standardized file using the service
            standardized_file = DataStandardizationService.find_standardized_file(filename, brand)
            if standardized_file and standardized_file.exists():
                file_path = standardized_file
                logger.info(f"Using standardized file: {file_path}")
            else:
                # Try direct lookup in standardized directory
                standardized_dir = brand_dirs["upload_dir"] / "standardized"
                original_name = Path(filename).stem
                standardized_filename = f"{original_name}_std.xlsx"
                potential_standardized = standardized_dir / standardized_filename
                
                if potential_standardized.exists():
                    file_path = potential_standardized
                    logger.info(f"Found standardized file: {file_path}")
                else:
                    # NO FALLBACK - Fail if standardized data not found
                    error_msg = f"Standardized data not found for {filename}. Expected file: {potential_standardized}"
                    logger.error(error_msg)
                    raise HTTPException(
                        status_code=404, 
                        detail=f"Standardized data not found. Please ensure standardized data has been created for {filename}. Expected location: {potential_standardized}"
                    )
        else:
            # Use original data - search in order: concat, raw, global fallbacks
            search_directories = [
                brand_dirs["concat_dir"], 
                brand_dirs["raw_dir"],
                settings.UPLOAD_DIR / "raw",  # Global upload directory as fallback
                settings.RAW_DIR  # Legacy raw directory as fallback
            ]
            file_path, _ = find_file_with_fallback(filename, search_directories)
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Read the data
        df = pd.read_excel(file_path) if file_path.suffix.lower() == '.xlsx' else pd.read_csv(file_path)
        
        # Validate variables exist in the dataset
        all_variables = [target_variable] + independent_variables
        missing_variables = [var for var in all_variables if var not in df.columns]
        if missing_variables:
            raise HTTPException(
                status_code=400,
                detail=f"Variables not found in dataset: {missing_variables}"
            )
        
        # Prepare data for modeling
        X = df[independent_variables].copy()
        y = df[target_variable].copy()
        
        # Log data information for debugging
        logger.info(f"Original data shape: {df.shape}")
        logger.info(f"Independent variables: {independent_variables}")
        logger.info(f"Target variable: {target_variable}")
        logger.info(f"X shape before cleaning: {X.shape}")
        logger.info(f"y shape before cleaning: {y.shape}")
        logger.info(f"X NaN count: {X.isnull().sum().sum()}")
        logger.info(f"y NaN count: {y.isnull().sum()}")
        
        # Handle missing values - ensure both X and y have no NaN values
        # Create a mask for rows where both features and target have no NaN values
        valid_mask = ~(X.isnull().any(axis=1) | y.isnull())
        
        X = X[valid_mask]
        y = y[valid_mask]
        
        logger.info(f"Valid rows after cleaning: {len(X)}")
        logger.info(f"X shape after cleaning: {X.shape}")
        logger.info(f"y shape after cleaning: {y.shape}")
        
        if len(X) == 0:
            raise HTTPException(
                status_code=400,
                detail="No valid data points after removing missing values"
            )
        
        # Additional validation to ensure no NaN values remain
        if X.isnull().any().any() or y.isnull().any():
            raise HTTPException(
                status_code=400,
                detail="Data still contains NaN values after cleaning"
            )
        
        # Train model based on type
        model_result = _train_model_by_type(
            X, y, model_type, model_parameters, validation_split, file_path, data_type
        )
        
        return {
            "success": True,
            "message": f"{model_type} model trained successfully",
            "data": model_result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in model training: {str(e)}")
        logger.error(f"Error training model: Model training failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Model training failed: {str(e)}")

@router.get("/list-models/{filename}")
async def list_trained_models(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup"),
    limit: int = Query(10, ge=1, le=100),
    offset: int = Query(0, ge=0)
) -> Dict[str, Any]:
    """
    List all trained models for a specific file
    
    Parameters:
    - filename: Name of the file to get models for
    - brand: Brand name for data lookup
    - limit: Maximum number of models to return
    - offset: Number of models to skip
    """
    try:
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Find the file
        search_directories = [brand_dirs["concat_dir"], brand_dirs["raw_dir"]]
        file_path, _ = find_file_with_fallback(filename, search_directories)
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Get models directory
        models_dir = file_path.parent / "models"
        if not models_dir.exists():
            return {
                "success": True,
                "data": {
                    "models": [],
                    "totalModels": 0,
                    "bestModel": None
                }
            }
        
        # Load model metadata
        models = []
        for model_file in models_dir.glob("*.json"):
            try:
                with open(model_file, 'r') as f:
                    model_data = json.load(f)
                    models.append(model_data)
            except Exception as e:
                logger.warning(f"Failed to load model {model_file}: {e}")
        
        # Sort by creation time (latest first)
        def get_sort_key(model):
            last_updated = model.get('trainingInfo', {}).get('lastUpdated', '')
            if last_updated:
                try:
                    from datetime import datetime
                    return datetime.fromisoformat(last_updated.replace('Z', '+00:00'))
                except:
                    return datetime.min
            return datetime.min
        
        models.sort(key=get_sort_key, reverse=True)
        
        # Log the sorted order for debugging
        logger.info(f"Models sorted by timestamp (latest first):")
        for i, model in enumerate(models):
            model_id = model.get('modelId', 'unknown')
            last_updated = model.get('trainingInfo', {}).get('lastUpdated', 'unknown')
            logger.info(f"  {i+1}. {model_id} - {last_updated}")
        
        # Apply pagination
        total_models = len(models)
        paginated_models = models[offset:offset + limit]
        
        # Get best model
        best_model = models[0] if models else None
        
        return {
            "success": True,
            "data": {
                "models": paginated_models,
                "totalModels": total_models,
                "bestModel": best_model
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing models: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to list models: {str(e)}")

@router.delete("/delete-model/{filename}/{model_id}")
async def delete_trained_model(
    filename: str,
    model_id: str,
    brand: str = Query(..., description="Brand name for data lookup")
) -> Dict[str, Any]:
    """
    Delete a trained model
    
    Parameters:
    - filename: Name of the file the model belongs to
    - model_id: ID of the model to delete
    - brand: Brand name for data lookup
    """
    try:
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Find the file
        search_directories = [brand_dirs["concat_dir"], brand_dirs["raw_dir"]]
        file_path, _ = find_file_with_fallback(filename, search_directories)
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Get models directory
        models_dir = file_path.parent / "models"
        if not models_dir.exists():
            raise HTTPException(status_code=404, detail="No models directory found")
        
        # Find and delete model file
        model_file = models_dir / f"{model_id}.json"
        if not model_file.exists():
            raise HTTPException(status_code=404, detail=f"Model {model_id} not found")
        
        # Delete the model file
        model_file.unlink()
        
        # Count remaining models
        remaining_models = len(list(models_dir.glob("*.json")))
        
        return {
            "success": True,
            "message": f"Model {model_id} deleted successfully",
            "data": {
                "deletedModelId": model_id,
                "remainingModels": remaining_models
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting model: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete model: {str(e)}")

def _train_model_by_type(X, y, model_type, model_parameters, validation_split, file_path, data_type="original"):
    """
    Train a model based on the specified type
    """
    try:
        from sklearn.model_selection import train_test_split
        from sklearn.linear_model import LinearRegression, Ridge
        from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
        from sklearn.preprocessing import StandardScaler
        import numpy as np
        
        # Final validation - ensure no NaN values in input data
        if X.isnull().any().any():
            raise ValueError("Input features (X) contain NaN values")
        if y.isnull().any():
            raise ValueError("Target variable (y) contains NaN values")
        
        logger.info(f"Training {model_type} model with {len(X)} samples and {len(X.columns)} features")
        
        # Split data
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=validation_split, random_state=42
        )
        
        # Scale features for ridge regression
        if model_type == "ridge":
            scaler = StandardScaler()
            X_train_scaled = scaler.fit_transform(X_train)
            X_test_scaled = scaler.transform(X_test)
        else:
            X_train_scaled = X_train
            X_test_scaled = X_test
        
        # Apply transformations based on model type
        if model_type == "log-linear":
            # Log transform target variable
            # Check for negative values before log transformation
            if (y_train < 0).any() or (y_test < 0).any():
                logger.warning("Negative values detected in target variable. Adding offset for log transformation.")
                min_val = min(y_train.min(), y_test.min())
                offset = abs(min_val) + 1
                y_train_transformed = np.log(y_train + offset)
                y_test_transformed = np.log(y_test + offset)
            else:
                y_train_transformed = np.log(y_train + 1)  # +1 to handle zeros
                y_test_transformed = np.log(y_test + 1)
            X_train_transformed = X_train_scaled
            X_test_transformed = X_test_scaled
        elif model_type == "log-log":
            # Log transform both features and target
            # Check for negative values before log transformation
            if (y_train < 0).any() or (y_test < 0).any():
                logger.warning("Negative values detected in target variable. Adding offset for log transformation.")
                min_val = min(y_train.min(), y_test.min())
                offset = abs(min_val) + 1
                y_train_transformed = np.log(y_train + offset)
                y_test_transformed = np.log(y_test + offset)
            else:
                y_train_transformed = np.log(y_train + 1)
                y_test_transformed = np.log(y_test + 1)
            
            # Check for negative values in features
            if (X_train_scaled < 0).any() or (X_test_scaled < 0).any():
                logger.warning("Negative values detected in features. Adding offset for log transformation.")
                min_val = min(X_train_scaled.min(), X_test_scaled.min())
                offset = abs(min_val) + 1
                X_train_transformed = np.log(X_train_scaled + offset)
                X_test_transformed = np.log(X_test_scaled + offset)
            else:
                X_train_transformed = np.log(X_train_scaled + 1)
                X_test_transformed = np.log(X_test_scaled + 1)
        else:
            y_train_transformed = y_train
            y_test_transformed = y_test
            X_train_transformed = X_train_scaled
            X_test_transformed = X_test_scaled
        
        # Select and train model
        if model_type == "ridge":
            alpha = model_parameters.get("alpha", 1.0)
            model = Ridge(alpha=alpha, random_state=42)
        elif model_type == "bayesian":
            # Simple Bayesian Ridge as approximation
            from sklearn.linear_model import BayesianRidge
            model = BayesianRidge(random_state=42)
        else:
            model = LinearRegression()
        
        # Check for potential singular matrix issues
        if len(X_train_transformed) < len(X.columns):
            raise ValueError(f"Not enough training samples ({len(X_train_transformed)}) for the number of features ({len(X.columns)}). Need at least {len(X.columns) + 1} samples.")
        
        # Check for insufficient degrees of freedom
        degrees_of_freedom = len(X_train_transformed) - len(X.columns) - 1
        if degrees_of_freedom <= 0:
            raise ValueError(f"Insufficient degrees of freedom ({degrees_of_freedom}) for reliable model training. You have {len(X_train_transformed)} samples and {len(X.columns)} features. Consider using fewer features or getting more data.")
        
        # Check for constant features (which can cause singular matrix)
        constant_features = []
        # Convert to numpy array if it's a DataFrame
        X_train_array = X_train_transformed.values if hasattr(X_train_transformed, 'values') else X_train_transformed
        for i, col in enumerate(X.columns):
            if np.std(X_train_array[:, i]) < 1e-10:
                constant_features.append(col)
        
        if constant_features:
            logger.warning(f"Constant features detected: {constant_features}. This may cause singular matrix issues.")
        
        # Train model with error handling
        start_time = datetime.now()
        try:
            model.fit(X_train_transformed, y_train_transformed)
        except np.linalg.LinAlgError as e:
            if "singular matrix" in str(e).lower():
                raise ValueError(f"Singular matrix error: The data may have perfect multicollinearity or insufficient variation. Try using Ridge regression or removing constant features. Original error: {str(e)}")
            else:
                raise ValueError(f"Linear algebra error during model training: {str(e)}")
        except Exception as e:
            raise ValueError(f"Model training failed: {str(e)}")
        
        training_time = (datetime.now() - start_time).total_seconds()
        
        # Make predictions
        y_pred = model.predict(X_test_transformed)
        
        # Transform predictions back if needed
        if model_type == "log-linear" or model_type == "log-log":
            y_pred_original = np.exp(y_pred) - 1
            y_test_original = np.exp(y_test_transformed) - 1
        else:
            y_pred_original = y_pred
            y_test_original = y_test_transformed
        
        # Calculate metrics
        r_squared = r2_score(y_test_original, y_pred_original)
        
        # Fix division by zero in adjusted R-squared calculation
        degrees_of_freedom = len(y_test) - len(X.columns) - 1
        if degrees_of_freedom > 0:
            adjusted_r_squared = 1 - (1 - r_squared) * (len(y_test) - 1) / degrees_of_freedom
        else:
            # If degrees of freedom is 0 or negative, set adjusted R-squared to R-squared
            logger.warning(f"Insufficient degrees of freedom ({degrees_of_freedom}) for adjusted R-squared calculation. Using R-squared instead.")
            adjusted_r_squared = r_squared
        
        rmse = np.sqrt(mean_squared_error(y_test_original, y_pred_original))
        mae = mean_absolute_error(y_test_original, y_pred_original)
        
        # Calculate variable statistics
        variables = []
        if hasattr(model, 'coef_'):
            for i, var_name in enumerate(X.columns):
                coefficient = model.coef_[i] if i < len(model.coef_) else 0
                
                # Calculate proper p-value using scipy
                try:
                    from scipy import stats
                    # Get residuals and calculate standard error
                    y_pred_var = model.predict(X_test_transformed)
                    residuals = y_test_transformed - y_pred_var
                    
                    # Fix division by zero in MSE calculation
                    degrees_of_freedom = len(y_test) - len(X.columns) - 1
                    if degrees_of_freedom > 0:
                        mse = np.sum(residuals ** 2) / degrees_of_freedom
                    else:
                        # If degrees of freedom is 0 or negative, use simple MSE
                        mse = np.mean(residuals ** 2)
                    
                    # Standard error of coefficient
                    X_with_intercept = np.column_stack([np.ones(len(X_test_transformed)), X_test_transformed])
                    try:
                        XtX_inv = np.linalg.inv(X_with_intercept.T @ X_with_intercept)
                        se_coef = np.sqrt(mse * XtX_inv[i+1, i+1]) if i+1 < XtX_inv.shape[0] else 0
                    except np.linalg.LinAlgError:
                        se_coef = 0
                    
                    # t-statistic and p-value
                    if se_coef > 0:
                        t_stat = coefficient / se_coef
                        # Calculate proper p-value using t-distribution
                        df = len(y_test) - len(X.columns) - 1
                        p_value = 2 * (1 - stats.t.cdf(abs(t_stat), df))
                        p_value = max(0, min(1, p_value))  # Ensure p-value is between 0 and 1
                    else:
                        t_stat = 0
                        p_value = 1.0
                        
                except ImportError:
                    # Fallback if scipy not available
                    t_stat = 0
                    p_value = 1.0
                
                # VIF calculation using statsmodels
                try:
                    from statsmodels.stats.outliers_influence import variance_inflation_factor
                    # Calculate VIF for this variable
                    # VIF requires the design matrix (X) and the index of the variable
                    vif = variance_inflation_factor(X_train_transformed, i)
                    # Handle infinite or very large VIF values
                    if np.isinf(vif) or vif > 1000:
                        vif = 1000.0  # Cap at 1000 for display purposes
                except Exception as e:
                    logger.warning(f"VIF calculation failed for variable {var_name}: {str(e)}")
                    vif = 1.0  # Fallback to 1.0 if calculation fails
                
                # Calculate elasticity @10% (percentage change in target for 10% change in variable)
                # Elasticity = (dY/Y) / (dX/X) = (dY/dX) * (X/Y) = coefficient * (X_mean / Y_mean)
                try:
                    # Get mean values for elasticity calculation from ORIGINAL data (not transformed)
                    # This is crucial for meaningful elasticity interpretation
                    X_mean = np.mean(X_train_original[:, i]) if hasattr(X_train_original, 'values') else np.mean(X_train_original[:, i])
                    y_mean = np.mean(y_train_original)
                    
                    # Calculate elasticity @10%: coefficient * (X_mean / Y_mean) * 0.1
                    # This gives the percentage change in target for a 10% change in the variable
                    if y_mean != 0 and X_mean != 0:
                        elasticity_10_percent = float(coefficient * (X_mean / y_mean) * 0.1)
                    else:
                        elasticity_10_percent = 0.0
                except Exception as e:
                    logger.warning(f"Elasticity calculation failed for variable {var_name}: {str(e)}")
                    elasticity_10_percent = 0.0
                
                variables.append({
                    "name": var_name,
                    "coefficient": float(coefficient),
                    "pValue": float(p_value),
                    "tStatistic": float(t_stat),
                    "vif": float(vif),
                    "elasticity10Percent": float(elasticity_10_percent),
                    "standardError": float(se_coef),
                    "confidenceInterval": [
                        float(coefficient - 1.96 * se_coef),
                        float(coefficient + 1.96 * se_coef)
                    ]
                })
        
        # Create model result
        model_id = f"{model_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        model_result = {
            "modelId": model_id,
            "modelName": f"{model_type.capitalize()} Model",
            "modelType": model_type,
            "dataType": data_type,  # Add data type information
            "rSquared": float(r_squared),
            "adjustedRSquared": float(adjusted_r_squared),
            "intercept": float(model.intercept_) if hasattr(model, 'intercept_') else 0,
            "variables": variables,
            "performanceMetrics": {
                "mape": float(np.mean(np.abs((y_test_original - y_pred_original) / np.where(y_test_original != 0, y_test_original, 1))) * 100) if len(y_test_original) > 0 else 0.0,
                "rmse": float(rmse),
                "mae": float(mae),
                "aic": 0,  # Would need more complex calculation
                "bic": 0   # Would need more complex calculation
            },
            "trainingInfo": {
                "trainingRows": len(X_train),
                "validationRows": len(X_test),
                "trainingTime": training_time,
                "lastUpdated": datetime.now().isoformat()
            }
        }
        
        # Save model metadata (file_path is defined above in the try block)
        _save_model_metadata(model_result, file_path)  # type: ignore
        
        return model_result
        
    except Exception as e:
        logger.error(f"Error in model training: {str(e)}")
        raise Exception(f"Model training failed: {str(e)}")

def _save_model_metadata(model_result, file_path):
    """
    Save model metadata to JSON file
    """
    try:
        models_dir = file_path.parent / "models"
        models_dir.mkdir(exist_ok=True)
        
        # Check if model with same parameters already exists
        existing_models = list(models_dir.glob("*.json"))
        for existing_model in existing_models:
            try:
                with open(existing_model, 'r') as f:
                    existing_data = json.load(f)
                    # Check if same model type and variables
                    if (existing_data.get('modelType') == model_result.get('modelType') and
                        existing_data.get('variables') == model_result.get('variables')):
                        logger.info(f"Model with same parameters already exists, skipping save")
                        return
            except Exception:
                continue
        
        model_file = models_dir / f"{model_result['modelId']}.json"
        with open(model_file, 'w') as f:
            json.dump(model_result, f, indent=2, default=str)
            
    except Exception as e:
        logger.warning(f"Failed to save model metadata: {e}")

# ========================================
# STATE MANAGEMENT ENDPOINTS
# ========================================

@router.post("/upload-data")
async def upload_data(
    file: UploadFile = File(...),
    brand_name: str = Form(...),
    analysis_name: str = Form(...)
):
    """
    Upload and process data for non-MMM analysis.
    """
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.csv')):
            raise HTTPException(
                status_code=400,
                detail="Only Excel (.xlsx) and CSV (.csv) files are supported"
            )
        
        # Read the uploaded file
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(await file.read()))
        else:
            df = pd.read_csv(io.BytesIO(await file.read()))
        
        # Get basic file info
        file_info = {
            "filename": file.filename,
            "brand_name": brand_name,
            "analysis_name": analysis_name,
            "upload_timestamp": datetime.now().isoformat(),
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "columns": df.columns.tolist(),
            "data_types": df.dtypes.astype(str).to_dict()
        }
        
        # Save the data to a temporary location for processing
        temp_filename = f"{brand_name}_{analysis_name}_{int(time.time())}.xlsx"
        temp_path = f"data/{brand_name}/data/{temp_filename}"
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        
        # Save the processed data
        df.to_excel(temp_path, index=False)
        
        # Update file info with saved path
        file_info["saved_path"] = temp_path
        
        return {
            "success": True,
            "message": "Data uploaded and processed successfully",
            "file_info": file_info
        }
        
    except Exception as e:
        logger.error(f"Error uploading data: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to upload and process data: {str(e)}"
        )

@router.post("/save-state")
async def save_state(state_data: dict):
    """
    Save the current state of the non-MMM analysis.
    """
    try:
        # Extract required information from state_data
        brand_name = state_data.get("brand_name")
        analysis_name = state_data.get("analysis_name")
        
        if not brand_name or not analysis_name:
            raise HTTPException(
                status_code=400,
                detail="Brand name and analysis name are required"
            )
        
        # Create state filename
        state_filename = f"{brand_name}_{analysis_name}_state.json"
        state_path = f"data/{brand_name}/data/{state_filename}"
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(state_path), exist_ok=True)
        
        # Save the state
        with open(state_path, 'w') as f:
            json.dump(state_data, f, indent=2)
        
        return {
            "success": True,
            "message": "State saved successfully",
            "state_path": state_path
        }
        
    except Exception as e:
        logger.error(f"Error saving state: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to save state: {str(e)}"
        )

@router.get("/load-state/{brand_name}/{analysis_name}")
async def load_state(brand_name: str, analysis_name: str):
    """
    Load the saved state for a non-MMM analysis.
    """
    try:
        # Create state filename
        state_filename = f"{brand_name}_{analysis_name}_state.json"
        state_path = f"data/{brand_name}/data/{state_filename}"
        
        # Check if state file exists
        if not os.path.exists(state_path):
            raise HTTPException(
                status_code=404,
                detail="State file not found"
            )
        
        # Load the state
        with open(state_path, 'r') as f:
            state_data = json.load(f)
        
        return {
            "success": True,
            "message": "State loaded successfully",
            "state_data": state_data
        }
        
    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail="State file not found"
        )
    except Exception as e:
        logger.error(f"Error loading state: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to load state: {str(e)}"
        )

# ========================================
# DATA STANDARDIZATION ENDPOINTS
# ========================================

@router.post("/create-standardized-data")
async def create_standardized_data(
    request_data: Dict[str, Any] = Body(...)
) -> Dict[str, Any]:
    """
    Create standardized data automatically when moving to model building step
    
    Request body should contain:
    - filename: Name of the original file
    - brand: Brand name for data lookup
    - analysisId: Analysis ID for file naming
    - method: Standardization method (default: 'zscore')
    """
    try:
        from app.services.data_standardization_service import DataStandardizationService
        
        filename = request_data.get("filename")
        brand = request_data.get("brand")
        analysis_id = request_data.get("analysisId")
        method = request_data.get("method", "zscore")
        
        if not all([filename, brand, analysis_id]):
            raise HTTPException(
                status_code=400,
                detail="filename, brand, and analysisId are required"
            )
        
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Find the original file (prioritize concatenated data if available)
        original_file_path = None
        search_directories = [
            brand_dirs["concat_dir"], 
            brand_dirs["raw_dir"],
            settings.UPLOAD_DIR / "raw",
            settings.RAW_DIR
        ]
        
        for directory in search_directories:
            if directory.exists():
                potential_file = directory / filename
                if potential_file.exists():
                    original_file_path = potential_file
                    break
        
        if not original_file_path:
            raise HTTPException(
                status_code=404,
                detail=f"Original file not found: {filename}"
            )
        
        # Check if standardized file already exists
        existing_standardized = DataStandardizationService.find_standardized_file(filename, brand, method)
        if existing_standardized:
            return {
                "success": True,
                "message": f"Standardized data already exists using {method} method",
                "data": {
                    "originalFile": str(original_file_path),
                    "standardizedFile": str(existing_standardized),
                    "standardizedFilename": existing_standardized.name,
                    "method": method,
                    "brand": brand,
                    "analysisId": analysis_id,
                    "alreadyExists": True
                }
            }
        
        # Create standardized file
        logger.info(f"🔄 Creating standardized file for {filename} using {method} method")
        logger.info(f"🔄 Original file: {original_file_path}")
        logger.info(f"🔄 Brand: {brand}, Analysis ID: {analysis_id}")
        
        standardized_file_path, metadata = DataStandardizationService.create_standardized_file(
            original_file_path=original_file_path,
            brand=brand,
            analysis_id=analysis_id,
            method=method
        )
        
        # Verify the file was created and is accessible
        if not standardized_file_path.exists():
            raise FileNotFoundError(f"Standardized file was not created: {standardized_file_path}")
        
        file_size = int(standardized_file_path.stat().st_size)  # Convert to Python int
        logger.info(f"✅ Standardized file created successfully: {standardized_file_path}")
        logger.info(f"✅ File size: {file_size} bytes")
        
        return {
            "success": True,
            "message": f"Standardized data created successfully using {method} method",
            "data": {
                "originalFile": str(original_file_path),
                "standardizedFile": str(standardized_file_path),
                "standardizedFilename": standardized_file_path.name,
                "method": method,
                "metadata": metadata,
                "brand": brand,
                "analysisId": analysis_id,
                "alreadyExists": False,
                "fileSize": file_size,
                "dataShape": metadata.get("standardized_shape", "unknown")
            }
        }
        
    except Exception as e:
        logger.error(f"Error creating standardized data: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create standardized data: {str(e)}"
        )

@router.post("/standardize-data")
async def standardize_data(
    request_data: Dict[str, Any] = Body(...)
) -> Dict[str, Any]:
    """
    Standardize data for Non-MMM analysis
    
    Request body should contain:
    - filename: Name of the original file
    - brand: Brand name for data lookup
    - analysisId: Analysis ID for file naming
    - method: Standardization method (zscore, minmax, robust, unit_vector)
    - columnsToStandardize: Optional list of specific columns to standardize
    - preserveColumns: Optional list of columns to preserve without standardization
    """
    try:
        from app.services.data_standardization_service import DataStandardizationService
        
        filename = request_data.get("filename")
        brand = request_data.get("brand")
        analysis_id = request_data.get("analysisId")
        method = request_data.get("method", "zscore")
        columns_to_standardize = request_data.get("columnsToStandardize")
        preserve_columns = request_data.get("preserveColumns")
        
        if not all([filename, brand, analysis_id]):
            raise HTTPException(
                status_code=400,
                detail="filename, brand, and analysisId are required"
            )
        
        # Validate standardization method
        valid_methods = DataStandardizationService.STANDARDIZATION_METHODS.keys()
        if method not in valid_methods:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid standardization method. Must be one of: {list(valid_methods)}"
            )
        
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Find the original file
        search_directories = [
            brand_dirs["concat_dir"], 
            brand_dirs["raw_dir"],
            settings.UPLOAD_DIR / "raw",  # Global upload directory as fallback
        ]
        
        from app.utils.file_utils import find_file_with_fallback
        original_file_path, _ = find_file_with_fallback(filename, search_directories)
        
        if not original_file_path:
            raise HTTPException(
                status_code=404, 
                detail=f"File {filename} not found for brand {brand}"
            )
        
        logger.info(f"Standardizing data for file: {original_file_path}")
        logger.info(f"Method: {method}, Brand: {brand}, Analysis ID: {analysis_id}")
        
        # Create standardized file
        standardized_file_path, metadata = DataStandardizationService.create_standardized_file(
            original_file_path=original_file_path,
            brand=brand,
            analysis_id=analysis_id,
            method=method,
            columns_to_standardize=columns_to_standardize,
            preserve_columns=preserve_columns
        )
        
        return {
            "success": True,
            "message": f"Data standardized successfully using {method} method",
            "data": {
                "originalFile": str(original_file_path),
                "standardizedFile": str(standardized_file_path),
                "standardizedFilename": standardized_file_path.name,
                "method": method,
                "metadata": metadata,
                "brand": brand,
                "analysisId": analysis_id
            }
        }
        
    except Exception as e:
        logger.error(f"Error in data standardization: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to standardize data: {str(e)}"
        )

@router.get("/standardization-status/{filename}")
async def get_standardization_status(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup"),
    analysis_id: Optional[str] = Query(None, description="Analysis ID to filter results")
) -> Dict[str, Any]:
    """
    Get standardization status for a file
    
    Returns information about available standardized versions of the file
    """
    try:
        from app.services.data_standardization_service import DataStandardizationService
        
        if not brand:
            raise HTTPException(status_code=400, detail="Brand parameter is required")
        
        # Find standardized files
        standardized_file = DataStandardizationService.find_standardized_file(
            filename, brand
        )
        
        if not standardized_file:
            return {
                "success": True,
                "data": {
                    "isStandardized": False,
                    "standardizedFiles": [],
                    "message": "No standardized version found"
                }
            }
        
        # Get metadata for all standardized versions
        metadata_list = DataStandardizationService.get_standardization_metadata(
            brand, analysis_id or "default"
        )
        
        # Filter metadata for this specific file
        file_metadata = []
        for metadata in metadata_list:
            if filename in metadata.get("original_file", ""):
                file_metadata.append(metadata)
        
        return {
            "success": True,
            "data": {
                "isStandardized": True,
                "standardizedFiles": [
                    {
                        "filename": metadata.get("standardized_file", "").split("/")[-1],
                        "method": metadata.get("method"),
                        "timestamp": metadata.get("timestamp"),
                        "columnsStandardized": metadata.get("columns_standardized", []),
                        "originalShape": metadata.get("original_shape"),
                        "standardizedShape": metadata.get("standardized_shape")
                    }
                    for metadata in file_metadata
                ],
                "latestStandardizedFile": standardized_file.name if standardized_file else None
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting standardization status: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to get standardization status: {str(e)}"
        )

@router.get("/download-standardized/{filename}")
async def download_standardized_file(
    filename: str,
    brand: str = Query(..., description="Brand name for data lookup"),
    method: Optional[str] = Query(None, description="Standardization method filter")
):
    """
    Download a standardized file
    
    Returns the standardized Excel file for download
    """
    try:
        from app.services.data_standardization_service import DataStandardizationService
        
        if not brand:
            raise HTTPException(status_code=400, detail="Brand parameter is required")
        
        # Find standardized file
        standardized_file = DataStandardizationService.find_standardized_file(
            filename, brand, method
        )
        
        if not standardized_file:
            raise HTTPException(
                status_code=404,
                detail=f"No standardized file found for {filename} with method {method or 'any'}"
            )
        
        # Return file for download
        from fastapi.responses import FileResponse
        return FileResponse(
            path=str(standardized_file),
            filename=standardized_file.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Error downloading standardized file: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to download standardized file: {str(e)}"
        )

@router.delete("/delete-column/{filename}")
async def delete_column(
    filename: str,
    column_name: str = Query(..., description="Name of the column to delete"),
    brand: str = Query(..., description="Brand name for data lookup")
) -> Dict[str, Any]:
    """
    Delete a specific column from the data file
    
    This endpoint removes a column from the Excel file and updates the file.
    The column will no longer appear in subsequent analysis steps.
    
    Parameters:
    - filename: Name of the file to modify
    - column_name: Name of the column to delete
    - brand: Brand name for data lookup
    """
    try:
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # CRITICAL: Always modify the RAW file, not concatenated or processed versions
        # This ensures that charts and other components read from the same file that gets modified
        search_directories = [
            brand_dirs["raw_dir"],  # ONLY search in raw directory first
            settings.UPLOAD_DIR / "raw",  # Global upload directory as fallback
            settings.RAW_DIR  # Legacy raw directory as fallback
        ]
        file_path, source_dir = find_file_with_fallback(filename, search_directories)
        if not file_path:
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        # Log which file we're modifying for debugging
        logger.info(f"Deleting column from file: {file_path} (found in: {source_dir})")
        
        # Read the Excel file
        df = pd.read_excel(file_path)
        
        # Check if column exists
        if column_name not in df.columns:
            raise HTTPException(
                status_code=404, 
                detail=f"Column '{column_name}' not found in file {filename}"
            )
        
        # Check if it's the target variable (prevent deletion)
        # We'll need to get this from the non-MMM state
        try:
            # Try to get target variable from non-MMM state
            nonmmm_states_dir = brand_dirs["metadata_dir"] / "nonmmm_states"
            if nonmmm_states_dir.exists():
                # Find the most recent non-MMM state file for this brand
                state_files = list(nonmmm_states_dir.glob("*.json"))
                if state_files:
                    # Sort by modification time to get the most recent
                    latest_state_file = max(state_files, key=lambda x: x.stat().st_mtime)
                    with open(latest_state_file, 'r') as f:
                        state_data = json.load(f)
                        target_variable = state_data.get('targetVariable')
                        if target_variable == column_name:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Cannot delete target variable '{column_name}'. Please select a different target variable first."
                            )
        except Exception as e:
            logger.warning(f"Could not check target variable: {str(e)}")
        
        # Remove the column
        df_modified = df.drop(columns=[column_name])
        
        # Save the modified file back
        df_modified.to_excel(file_path, index=False)
        
        # Clean up any cached data summaries that might contain the deleted column
        try:
            _cleanup_cached_data_summaries(brand, filename, column_name)
        except Exception as cleanup_error:
            logger.warning(f"Failed to cleanup cached data summaries: {cleanup_error}")
        
        # Get updated column information
        remaining_columns = df_modified.columns.tolist()
        
        logger.info(f"Successfully deleted column '{column_name}' from {filename}")
        
        result = {
            "success": True,
            "message": f"Column '{column_name}' deleted successfully",
            "data": {
                "deletedColumn": column_name,
                "remainingColumns": remaining_columns,
                "totalColumns": len(remaining_columns),
                "filePath": str(file_path)
            }
        }
        
        logger.info(f"Returning delete result: {result}")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting column: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to delete column: {str(e)}"
        )


def _cleanup_cached_data_summaries(brand: str, filename: str, deleted_column: str):
    """
    Clean up cached data summaries that might contain the deleted column
    
    This function removes any cached data summaries that contain the deleted column
    to ensure that subsequent data-summary calls return fresh data.
    
    Args:
        brand: Brand name for data lookup
        filename: Name of the file that was modified
        deleted_column: Name of the column that was deleted
    """
    try:
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Look for cached data summaries in the nonmmm_summaries directory
        summaries_dir = brand_dirs["metadata_dir"] / "nonmmm_summaries"
        if summaries_dir.exists():
            # Find all JSON files in the summaries directory
            summary_files = list(summaries_dir.glob("*.json"))
            
            for summary_file in summary_files:
                try:
                    # Read the cached summary
                    with open(summary_file, 'r', encoding='utf-8') as f:
                        summary_data = json.load(f)
                    
                    # Check if this summary is for the same file and contains the deleted column
                    if (summary_data.get('filename') == filename and 
                        'variables' in summary_data and 
                        any(var.get('name') == deleted_column for var in summary_data['variables'])):
                        
                        # Remove the cached summary file
                        summary_file.unlink()
                        logger.info(f"Removed cached data summary: {summary_file.name}")
                        
                except Exception as file_error:
                    logger.warning(f"Failed to process cached summary {summary_file.name}: {file_error}")
        
        # Also clean up any stored data summaries that might contain the deleted column
        # This covers the case where summaries are stored via the store-summary endpoint
        stored_summaries_dir = brand_dirs["metadata_dir"] / "stored_summaries"
        if stored_summaries_dir.exists():
            stored_files = list(stored_summaries_dir.glob("*.json"))
            
            for stored_file in stored_files:
                try:
                    with open(stored_file, 'r', encoding='utf-8') as f:
                        stored_data = json.load(f)
                    
                    # Check if this stored summary contains the deleted column
                    if (stored_data.get('filename') == filename and 
                        'data' in stored_data and 
                        'variables' in stored_data['data'] and
                        any(var.get('name') == deleted_column for var in stored_data['data']['variables'])):
                        
                        # Remove the stored summary file
                        stored_file.unlink()
                        logger.info(f"Removed stored data summary: {stored_file.name}")
                        
                except Exception as file_error:
                    logger.warning(f"Failed to process stored summary {stored_file.name}: {file_error}")
        
        logger.info(f"Cache cleanup completed for deleted column '{deleted_column}' in file '{filename}'")
        
    except Exception as e:
        logger.error(f"Error during cache cleanup: {str(e)}")
        # Don't raise the exception as this is a cleanup operation
        # The main deletion operation should still succeed


@router.delete("/delete-variable/{filename}")
async def delete_variable(
    filename: str,
    brand: str = Query(..., description="Brand name"),
    column: str = Query(..., description="Column name to delete")
):
    """
    Delete a variable/column from the Excel file
    
    Args:
        filename: Name of the uploaded file
        brand: Brand name for file organization
        column: Name of the column to delete
        
    Returns:
        Success message confirming deletion
    """
    try:
        logger.info(f"Deleting variable '{column}' from {filename} (brand: {brand})")
        
        # Get brand directories
        brand_dirs = settings.get_brand_directories(brand)
        
        # Find the file using the same search strategy as other endpoints
        search_directories = [
            brand_dirs["raw_dir"],  # Read from raw file first
            brand_dirs["concat_dir"],  # Fallback to concatenated file
            settings.UPLOAD_DIR / "raw",  # Global upload directory as fallback
            settings.RAW_DIR  # Legacy raw directory as fallback
        ]
        file_path, source_dir = find_file_with_fallback(filename, search_directories)
        
        if not file_path or not file_path.exists():
            raise HTTPException(status_code=404, detail=f"File {filename} not found for brand {brand}")
        
        logger.info(f"Found file for deletion: {file_path} (source: {source_dir})")
        
        # Load the data
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        # Check if column exists
        if column not in df.columns:
            raise HTTPException(status_code=404, detail=f"Column '{column}' not found in file")
        
        # Check if it's the target variable (prevent deletion)
        try:
            # Try to get target variable from non-MMM state
            nonmmm_states_dir = brand_dirs["metadata_dir"] / "nonmmm_states"
            if nonmmm_states_dir.exists():
                # Find the most recent non-MMM state file for this brand
                state_files = list(nonmmm_states_dir.glob("*.json"))
                if state_files:
                    # Sort by modification time to get the most recent
                    latest_state_file = max(state_files, key=lambda x: x.stat().st_mtime)
                    with open(latest_state_file, 'r') as f:
                        state_data = json.load(f)
                        target_variable = state_data.get('targetVariable')
                        if target_variable == column:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Cannot delete target variable '{column}'. Please select a different target variable first."
                            )
        except Exception as e:
            logger.warning(f"Could not check target variable: {str(e)}")
        
        # Delete the column
        df = df.drop(columns=[column])
        
        # Save the modified data back to the file
        if file_path.suffix.lower() == '.xlsx':
            # For Excel files, save to the same file
            df.to_excel(file_path, index=False)
        elif file_path.suffix.lower() == '.csv':
            # For CSV files, save to the same file
            df.to_csv(file_path, index=False)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format for deletion")
        
        # Clean up cached data summaries that might contain the deleted column
        try:
            _cleanup_cached_data_summaries(brand, filename, column)
        except Exception as cleanup_error:
            logger.warning(f"Failed to cleanup cached data summaries: {cleanup_error}")
        
        logger.info(f"Successfully deleted variable '{column}' from {filename}")
        
        return BaseResponse(
            success=True,
            message=f"Variable '{column}' deleted successfully",
            data={
                "deletedColumn": column,
                "remainingColumns": len(df.columns),
                "remainingRows": len(df)
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting variable: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting variable: {str(e)}")
